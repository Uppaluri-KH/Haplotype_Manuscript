import openpyxl

# Load the existing Excel file
input_excel_file = 'Gene_allele_definition_table.xlsx'
workbook = openpyxl.load_workbook(input_excel_file)

# Select the "Alleles" sheet
sheet_name = "Alleles"
sheet = workbook[sheet_name]

# Starting from the 9th row in column A
start_row = 9

# Create a new sheet to store the duplicated values
output_sheet_name = "DuplicatedValues"
output_sheet = workbook.create_sheet(title=output_sheet_name)

# Write headers to the new sheet
output_sheet.append(["Haplotype", "ALT", "Genomic_pos", "rsID"])

# Iterate over values in column A from the 9th row onwards
for row_number in range(start_row, sheet.max_row + 1):
    cell_value_A = sheet[f'A{row_number}'].value
    
    # Check if any of the values in columns B, C, D, etc. are null
    if any(sheet.cell(row=row_number, column=col).value is None for col in range(2, sheet.max_column + 1)):
        # Get non-None values from columns B, C, D, etc.
        values_from_columns = [
            sheet.cell(row=row_number, column=col).value
            for col in range(2, sheet.max_column + 1)
            if sheet.cell(row=row_number, column=col).value is not None
        ]
        
        # Get the respective 4th and 6th row values only for non-null values in column A
        if cell_value_A is not None:
            # Filter out the 4th and 6th row values based on non-null values in column A
            fourth_row_values = [
                sheet.cell(row=4, column=col).value
                for col in range(2, sheet.max_column + 1)
                if sheet.cell(row=row_number, column=col).value is not None
            ]
            sixth_row_values = [
                sheet.cell(row=6, column=col).value
                for col in range(2, sheet.max_column + 1)
                if sheet.cell(row=row_number, column=col).value is not None
            ]

            # Append each value in a new row
            for individual_value, fourth_value, sixth_value in zip(values_from_columns, fourth_row_values, sixth_row_values):
                output_sheet.append([cell_value_A, individual_value, fourth_value, sixth_value])

# Save the workbook with the new sheet
output_excel_file = 'Gene_allele_definition.xlsx'
workbook.save(output_excel_file)

print(f"Values saved to the new Excel file: {output_excel_file}")